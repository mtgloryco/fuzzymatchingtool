import openpyxl as op
import pandas as pd
from rapidfuzz import process, fuzz
from tkinter import filedialog, messagebox, StringVar, simpledialog, ttk, scrolledtext
import tkinter as tk
from datetime import datetime
import openpyxl.styles
import collections
import ctypes
import time
import re
import threading
from rounded_button import round_button
from openpyxl import load_workbook

# Set DPI awareness for better display on Windows
try:
    ctypes.windll.shcore.SetProcessDpiAwareness(1)
except:
    pass

# Global variable declarations
progress_label = None

# ---------------------------- Utility Functions ---------------------------- #

def get_sheet_names(file_path):
    """Get all sheet names from an Excel file."""
    try:
        workbook = op.load_workbook(file_path, read_only=True)
        return workbook.sheetnames
    except Exception as e:
        messagebox.showerror("Error", f"Could not read sheets: {e}")
        return ["Sheet1"]

def get_tables_in_sheet(file_path, sheet_name):
    """Return a list of real Excel tables (ListObjects) in the given sheet."""
    try:
        wb = load_workbook(filename=file_path)
        sheet = wb[sheet_name]
        tables = []
        for table in sheet.tables.values():
            tables.append(table.name)
        if not tables:
            tables.append("Full Sheet")
        return tables
    except Exception as e:
        messagebox.showerror("Error", f"Could not read tables: {e}")
        return ["Full Sheet"]

def get_columns_in_table(file_path, sheet_name, table_name):
    """Get columns from a specific table or the entire sheet."""
    try:
        if table_name == "Full Sheet":
            df = pd.read_excel(file_path, sheet_name)
            return list(df.columns)
        elif table_name.startswith("Table:"):
            pure_table_name = table_name.replace("Table: ", "").replace(" ++", "")
            
            workbook = op.load_workbook(file_path)
            sheet = workbook[sheet_name]
            
            table_range = None
            
            if hasattr(sheet, '_tables'):
                for table in sheet._tables:
                    table_display_name = getattr(table, 'displayName', None)
                    if table_display_name == pure_table_name:
                        table_range = table.ref
                        break
            
            if not table_range and hasattr(workbook, 'defined_names'):
                if pure_table_name in workbook.defined_names:
                    for sheet_title, cell_range in workbook.defined_names[pure_table_name].destinations:
                        if sheet_title == sheet_name:
                            table_range = cell_range
                            break
            
            if table_range:
                if ':' in table_range:
                    start_cell, end_cell = table_range.split(':')
                    cell_range = sheet[table_range]
                    if isinstance(cell_range, tuple) and len(cell_range) > 0:
                        header_row = cell_range[0]
                        return [cell.value for cell in header_row if cell.value]
                
                df = pd.read_excel(file_path, sheet_name)
                return list(df.columns)
            else:
                df = pd.read_excel(file_path, sheet_name)
                return list(df.columns)
    except Exception as e:
        messagebox.showerror("Error", f"Could not read columns: {e}")
        return []

# ---------------------------- UI Update Functions ---------------------------- #

def update_sheet_dropdown(file_path, is_file1=True):
    """Update the sheet dropdown with sheet names from the selected file."""
    try:
        sheet_names = get_sheet_names(file_path)
        
        if is_file1:
            sheet1_dropdown['menu'].delete(0, 'end')
            selected_sheet1.set(sheet_names[0])
            
            for sheet in sheet_names:
                sheet1_dropdown['menu'].add_command(label=sheet, 
                                                  command=lambda s=sheet: selected_sheet1.set(s))
            update_table_dropdown(file_path, selected_sheet1.get(), is_file1=True)
        else:
            sheet2_dropdown['menu'].delete(0, 'end')
            selected_sheet2.set(sheet_names[0])
            
            for sheet in sheet_names:
                sheet2_dropdown['menu'].add_command(label=sheet, 
                                                  command=lambda s=sheet: selected_sheet2.set(s))
            update_table_dropdown(file_path, selected_sheet2.get(), is_file1=False)
    except Exception as e:
        messagebox.showerror("Error", f"Could not read sheets: {e}")

def update_table_dropdown(file_path, sheet_name, is_file1=True):
    """Update the table dropdown with tables from the selected sheet."""
    try:
        table_names = get_tables_in_sheet(file_path, sheet_name)
        
        if is_file1:
            table1_dropdown['menu'].delete(0, 'end')
            selected_table1.set(table_names[0])
            
            for table in table_names:
                table1_dropdown['menu'].add_command(label=table, 
                                                  command=lambda t=table: selected_table1.set(t))
            update_column_dropdown(file_path, sheet_name, selected_table1.get(), is_file1=True)
        else:
            table2_dropdown['menu'].delete(0, 'end')
            selected_table2.set(table_names[0])
            
            for table in table_names:
                table2_dropdown['menu'].add_command(label=table, 
                                                  command=lambda t=table: selected_table2.set(t))
            update_column_dropdown(file_path, sheet_name, selected_table2.get(), is_file1=False)
    except Exception as e:
        messagebox.showerror("Error", f"Could not read tables: {e}")

def update_column_dropdown(file_path, sheet_name, table_name, is_file1=True):
    """Update the column dropdown with columns from the selected table."""
    try:
        columns = get_columns_in_table(file_path, sheet_name, table_name)
        
        if is_file1:
            column1_dropdown['menu'].delete(0, 'end')
            if columns:
                selected_col1.set(columns[0])
                for col in columns:
                    column1_dropdown['menu'].add_command(label=col, 
                                                     command=lambda c=col: selected_col1.set(c))
                status_label.config(text=f"Ready - File 1 column set to: {selected_col1.get()}")
            else:
                selected_col1.set("")
                status_label.config(text="No columns found in selected table")
                
            load_table_data(file_path, sheet_name, table_name, tree1, "File 1")
            
        else:
            column2_dropdown['menu'].delete(0, 'end')
            if columns:
                selected_col2.set(columns[0])
                for col in columns:
                    column2_dropdown['menu'].add_command(label=col, 
                                                     command=lambda c=col: selected_col2.set(c))
                status_label.config(text=f"Ready - File 2 column set to: {selected_col2.get()}")
            else:
                selected_col2.set("")
                status_label.config(text="No columns found in selected table")
                
            load_table_data(file_path, sheet_name, table_name, tree2, "File 2")
            
    except Exception as e:
        messagebox.showerror("Error", f"Could not update columns: {e}")

def auto_adjust_columns(tree):
    """Auto-adjust column widths based on content"""
    for col in tree["columns"]:
        max_len = max(
            len(str(tree.heading(col)["text"])),
            *[len(str(tree.set(item, col))) for item in tree.get_children()],
            10
        )
        tree.column(col, width=max_len * 8)

def load_table_data(file_path, sheet_name, table_name, tree, file_label):
    """Load only the selected table's data into the treeview."""
    try:
        if table_name == "Full Sheet":
            df = pd.read_excel(file_path, sheet_name)
        else:
            wb = load_workbook(filename=file_path)
            ws = wb[sheet_name]
            table = ws.tables[table_name]
            data = ws[table.ref]
            rows = [[cell.value for cell in row] for row in data]
            df = pd.DataFrame(rows[1:], columns=rows[0])
        
        tree.delete(*tree.get_children())
        tree["columns"] = list(df.columns)
        
        for col in df.columns:
            tree.heading(col, text=col)
            tree.column(col, width=100)
            
        for _, row in df.iterrows():
            tree.insert("", "end", values=list(row))
            
        tree.after(100, lambda: auto_adjust_columns(tree))
        
    except Exception as e:
        messagebox.showerror("Error", f"Could not load table data: {e}")

# ---------------------------- Event Handlers ---------------------------- #

def on_sheet1_selected(*args):
    if hasattr(root, 'file1_path') and root.file1_path:
        update_table_dropdown(root.file1_path, selected_sheet1.get(), is_file1=True)
        try:
            workbook = op.load_workbook(root.file1_path)
            workbook.active = workbook[selected_sheet1.get()]
            workbook.save(root.file1_path)
        except Exception as e:
            print(f"Could not set active sheet in file: {e}")

def on_sheet2_selected(*args):
    if hasattr(root, 'file2_path') and root.file2_path:
        update_table_dropdown(root.file2_path, selected_sheet2.get(), is_file1=False)
        try:
            workbook = op.load_workbook(root.file2_path)
            workbook.active = workbook[selected_sheet2.get()]
            workbook.save(root.file2_path)
        except Exception as e:
            print(f"Could not set active sheet in file: {e}")

def on_table1_selected(*args):
    if hasattr(root, 'file1_path') and root.file1_path:
        load_table_data(root.file1_path, selected_sheet1.get(), selected_table1.get(), tree1, "File 1")
        update_column_dropdown(root.file1_path, selected_sheet1.get(), selected_table1.get(), is_file1=True)

def on_table2_selected(*args):
    if hasattr(root, 'file2_path') and root.file2_path:
        load_table_data(root.file2_path, selected_sheet2.get(), selected_table2.get(), tree2, "File 2")
        update_column_dropdown(root.file2_path, selected_sheet2.get(), selected_table2.get(), is_file1=False)

def on_column_click(event):
    """Handle column click for File 1."""
    column_id = event.widget.identify_column(event.x)
    
    if column_id:
        try:
            column_index = int(column_id.strip('#')) - 1
            if column_index >= 0:
                column_name = tree1.heading(column_id)['text']
                if column_name:
                    selected_col1.set(column_name)
                    status_label.config(text=f"File 1 column selected: {column_name}")
                    
                    for col_id in tree1["columns"]:
                        col_idx = tree1["columns"].index(col_id)
                        tree1.tag_configure(f"col_{col_idx}", background="white")
                    
                    tree1.tag_configure(f"col_{column_index}", background="light gray")
                    
                    for item_id in tree1.get_children():
                        tree1.item(item_id, tags=(f"col_{column_index}",))
                    
                    style_column_cells(tree1, column_index)
        except Exception as e:
            print(f"Error in column selection: {e}")

def on_column_click_2(event):
    """Handle column click for File 2."""
    column_id = event.widget.identify_column(event.x)
    if column_id:
        try:
            column_index = int(column_id.strip('#')) - 1
            if column_index >= 0:
                column_name = tree2.heading(column_id)['text']
                if column_name:
                    selected_col2.set(column_name)
                    status_label.config(text=f"File 2 column selected: {column_name}")
                    
                    for col_id in tree2["columns"]:
                        col_idx = tree2["columns"].index(col_id)
                        tree2.tag_configure(f"col_{col_idx}", background="white")
                    
                    tree2.tag_configure(f"col_{column_index}", background="light gray")
                    
                    for item_id in tree2.get_children():
                        tree2.item(item_id, tags=(f"col_{column_index}",))
                    
                    style_column_cells(tree2, column_index)
        except Exception as e:
            print(f"Error in column selection: {e}")

def style_column_cells(tree, column_index):
    """Style individual cells in the selected column."""
    for i in range(len(tree["columns"])):
        tag_name = f"col_{i}"
        tree.tag_configure(tag_name, background="white")
    
    tag_name = f"col_{column_index}"
    tree.tag_configure(tag_name, background="light gray")
    
    style = ttk.Style()
    style.map("Treeview", 
              foreground=fixed_map("foreground"),
              background=fixed_map("background"))

def fixed_map(option):
    # Fix for setting text colour for Tkinter ttk Treeview
    return [elm for elm in style.map("Treeview", query_opt=option)
            if elm[:2] != ("!disabled", "!selected")]

# ---------------------------- File Operations ---------------------------- #

def select_file1():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if file_path:
        root.file1_path = file_path
        filename = file_path.split('/')[-1]
        frame1.config(text=filename)
        update_sheet_dropdown(file_path, is_file1=True)

def select_file2():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if file_path:
        root.file2_path = file_path
        filename = file_path.split('/')[-1]
        frame2.config(text=filename)
        update_sheet_dropdown(file_path, is_file1=False)

# ---------------------------- Matching Functions ---------------------------- #

def show_progress_ui():
    global progress_label  # Make progress_label global
    progress_window = tk.Toplevel(root)
    progress_window.title("Matching Progress")
    progress_window.geometry("300x150")
    progress_window.transient(root)
    progress_window.grab_set()
    
    progress_window.geometry("+%d+%d" % (
        root.winfo_rootx() + root.winfo_width()/2 - 150,
        root.winfo_rooty() + root.winfo_height()/2 - 75))

    progress_label = ttk.Label(progress_window, text="Progress: 0%")
    progress_label.pack(pady=10)

    progress_bar = ttk.Progressbar(
        progress_window,
        variable=progress_var,
        maximum=100,
        mode='determinate',
        length=200
    )
    progress_bar.pack(pady=10)
    
    return progress_window

def show_advanced_options():
    """Toggle display of advanced options"""
    if advanced_options_frame.winfo_ismapped():
        advanced_options_frame.grid_remove()
        advanced_toggle_btn.config(text="Show Advanced Options")
    else:
        advanced_options_frame.grid()
        advanced_toggle_btn.config(text="Hide Advanced Options")

def show_stats(stats, result_filename, threshold):
    """Show statistics in a scrollable dialog"""
    stats_window = tk.Toplevel(root)
    stats_window.title("Matching Statistics")
    stats_window.geometry("500x400")
    
    text_area = scrolledtext.ScrolledText(
        stats_window, 
        wrap=tk.WORD,
        width=60,
        height=20,
        font=("Consolas", 10)
    )
    text_area.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)
    
    message = (
        f"Matching complete!\n\n"
        f"Results saved as '{result_filename}'\n\n"
        f"Statistics:\n"
        f"- Total compared: {stats['total']}\n"
        f"- Perfect matches: {stats['perfect_matches']}\n"
        f"- Above threshold ({threshold}%): {stats['above_threshold']}\n"
        f"- Near threshold: {stats['near_threshold']}\n"
        f"- Below threshold: {stats['below_threshold']}\n\n"
        f"Sheet 1: Matches >= {threshold}%\n"
        "Sheet 2: All matches\n"
        "Sheet 3: Detailed statistics"
    )
    
    text_area.insert(tk.INSERT, message)
    text_area.configure(state='disabled')
    
    close_btn = ttk.Button(
        stats_window,
        text="Close",
        command=stats_window.destroy
    )
    close_btn.pack(pady=5)
    
    stats_window.transient(root)
    stats_window.grab_set()
    root.wait_window(stats_window)

def preprocess_value(value, trim_whitespace, normalize_case, ignore_punct):
    """Preprocess a value based on the selected options"""
    if pd.isna(value):
        return ""
    s = str(value)
    if trim_whitespace:
        s = s.strip()
    if normalize_case:
        s = s.lower()
    if ignore_punct:
        s = re.sub(r'[^\w\s]', '', s)
    return s

def perform_matching():
    """Main matching function to run in a separate thread"""
    try:
        # Get matching options
        threshold = float(threshold_var.get())
        case_sensitive = case_sensitive_var.get()
        ignore_punct = ignore_punct_var.get()
        trim_whitespace = trim_whitespace_var.get()
        normalize_case = normalize_case_var.get() and not case_sensitive
        algorithm = algorithm_var.get()
        
        # Load data
        if selected_table1.get() == "Full Sheet":
            df1 = pd.read_excel(root.file1_path, selected_sheet1.get())
        else:
            df1 = pd.read_excel(root.file1_path, selected_sheet1.get())
            
        if selected_table2.get() == "Full Sheet":
            df2 = pd.read_excel(root.file2_path, selected_sheet2.get())
        else:
            df2 = pd.read_excel(root.file2_path, selected_sheet2.get())

        # Validate columns
        df1.columns = [str(col).strip() for col in df1.columns]
        df2.columns = [str(col).strip() for col in df2.columns]

        col1 = selected_col1.get().strip()
        col2 = selected_col2.get().strip()

        if col1 not in df1.columns:
            col1_match = next((col for col in df1.columns if col.lower() == col1.lower()), None)
            if col1_match:
                col1 = col1_match
            else:
                raise ValueError(f"Selected column '{selected_col1.get()}' not found in File 1")
        
        if col2 not in df2.columns:
            col2_match = next((col for col in df2.columns if col.lower() == col2.lower()), None)
            if col2_match:
                col2 = col2_match
            else:
                raise ValueError(f"Selected column '{selected_col2.get()}' not found in File 2")

        # Preprocess all values to match
        values_to_match = [
            preprocess_value(v, trim_whitespace, normalize_case, ignore_punct) 
            for v in df2[col2]
        ]
        
        # Choose scoring algorithm
        if algorithm == "Jaro-Winkler":
            scorer = fuzz.WRatio
        elif algorithm == "Ratio":
            scorer = fuzz.ratio
        else:  # Levenshtein
            scorer = None

        # Prepare results
        results = []
        total_rows = len(df1)
        stats = {
            'total': 0,
            'above_threshold': 0,
            'near_threshold': 0,
            'below_threshold': 0,
            'perfect_matches': 0
        }

        # Process each row
        for i, row in df1.iterrows():
            value = preprocess_value(row[col1], trim_whitespace, normalize_case, ignore_punct)
            
            if value:
                if scorer:
                    matches = process.extract(
                        value, 
                        values_to_match, 
                        scorer=scorer,
                        limit=1
                    )
                else:
                    matches = process.extract(
                        value, 
                        values_to_match, 
                        limit=1
                    )
                
                if matches:
                    match_value, match_score, _ = matches[0]
                    stats['total'] += 1
                    
                    if match_score == 100:
                        stats['perfect_matches'] += 1
                    if match_score >= threshold:
                        stats['above_threshold'] += 1
                    elif match_score >= threshold * 0.8:
                        stats['near_threshold'] += 1
                    else:
                        stats['below_threshold'] += 1
                        
                    results.append({
                        'source_value': row[col1],
                        'matched_value': match_value,
                        'score': match_score,
                        
                    })
            
            # Update progress
            progress = (i + 1) / total_rows * 100
            progress_var.set(progress)
            progress_label.config(
                text=f"Processing: {i+1}/{total_rows} rows\n"
                     f"Matches: {stats['above_threshold']} above threshold"
            )
            root.update()

        # Generate output
        all_results_df = pd.DataFrame(results)
        filtered_df = all_results_df[all_results_df['score'] >= threshold]

        # Save results
        if result_filename.endswith('.csv'):
            filtered_df.to_csv(result_filename, index=False)
            all_results_filename = result_filename.replace('.csv', '_all_results.csv')
            all_results_df.to_csv(all_results_filename, index=False)
        else:
            with pd.ExcelWriter(result_filename) as writer:
                filtered_df.to_excel(writer, sheet_name="Filtered Results", index=False)
                all_results_df.to_excel(writer, sheet_name="All Results", index=False)
                
                # Add stats sheet
                stats_df = pd.DataFrame.from_dict(stats, orient='index', columns=['Count'])
                stats_df.to_excel(writer, sheet_name="Statistics")

        # Show completion
        progress_var.set(100)
        progress_label.config(text="Done!")
        root.update()
        time.sleep(0.5)
        progress_window.destroy()
        
        # Show stats in main thread
        root.after(0, lambda: show_stats(stats, result_filename, threshold))
        status_label.config(text=f"Matching complete. {stats['above_threshold']} matches found.")

    except Exception as e:
        progress_window.destroy()
        messagebox.showerror("Error", f"An error occurred: {e}")
        status_label.config(text=f"Error: {str(e)}")

def start_matching():
    """Start the matching process with thread handling"""
    try:
        global result_filename, progress_window
        
        # Get save filename
        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        initial_filename = f"fuzzy_match_{current_time}.xlsx"
        result_filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("CSV Files", "*.csv")],
            initialfile=initial_filename,
            title="Save Match Results As"
        )
        
        if not result_filename:
            return

        # Show progress window
        progress_window = show_progress_ui()
        progress_var.set(0)
        progress_label.config(text="Initializing...")
        root.update()
        
        status_label.config(text="Processing...")
        root.update()
        
        # Validation checks
        if not hasattr(root, 'file1_path') or not hasattr(root, 'file2_path'):
            raise ValueError("Please select both Excel files first")

        if not selected_col1.get() or not selected_col2.get():
            raise ValueError("Please select columns for matching from both files")

        # Start the matching thread
        threading.Thread(target=perform_matching, daemon=True).start()

    except Exception as e:
        if 'progress_window' in locals():
            progress_window.destroy()
        messagebox.showerror("Error", f"An error occurred: {e}")
        status_label.config(text=f"Error: {str(e)}")

# ---------------------------- Main UI Setup ---------------------------- #

# Create main window
root = tk.Tk()
root.title("Excel Fuzzy Matching Tool")
root.geometry("1400x800")

# Apply style
style = ttk.Style()
style.theme_use("vista")
style.configure("Treeview.Heading", foreground="#7b6cd9", font=('Helvetica', 10, 'bold'))

# Create main horizontal frame
main_horizontal_frame = ttk.Frame(root)
main_horizontal_frame.pack(fill="both", expand=True)

# ---------------------------- Left Panel (File Selection) ---------------------------- #

sheet_frame = ttk.LabelFrame(main_horizontal_frame, text="File Selection", width=250)
sheet_frame.pack(side="left", fill="y", padx=5, pady=5)
sheet_frame.pack_propagate(False)

# File 1 button
file1_btn = round_button(sheet_frame, text="Choose file 1", fill="#7b6cd9", radius=25, 
                        command=select_file1, font=('Poppins', 9, 'bold'))
file1_btn.grid(row=1, column=0, pady=5, padx=5, sticky="ew")

# Separator
ttk.Separator(sheet_frame, orient='horizontal').grid(row=2, column=0, sticky='ew', pady=10)

# File 2 button
file2_btn = round_button(sheet_frame, text="Choose file 2", fill="#7b6cd9", radius=25, 
                        command=select_file2, font=('Poppins', 9, 'bold'))
file2_btn.grid(row=4, column=0, pady=5, padx=5, sticky="ew")

# ---------------------------- Center Panel (Data Display) ---------------------------- #

files_frame = ttk.LabelFrame(main_horizontal_frame, text="Excel Data")
files_frame.pack(side="left", fill="both", expand=True, padx=5, pady=5)

# Canvas and scrollbars for files frame
files_canvas = tk.Canvas(files_frame)
files_scrollbar_y = ttk.Scrollbar(files_frame, orient="vertical", command=files_canvas.yview)
files_scrollbar_x = ttk.Scrollbar(files_frame, orient="horizontal", command=files_canvas.xview)
files_scroll_frame = ttk.Frame(files_canvas)

files_scroll_frame.bind("<Configure>", lambda e: files_canvas.configure(scrollregion=files_canvas.bbox("all")))
files_canvas.create_window((0, 0), window=files_scroll_frame, anchor="nw")
files_canvas.configure(yscrollcommand=files_scrollbar_y.set, xscrollcommand=files_scrollbar_x.set)

files_scrollbar_y.pack(side="right", fill="y")
files_scrollbar_x.pack(side="bottom", fill="x")
files_canvas.pack(side="left", fill="both", expand=True)

# Mouse wheel bindings
def _on_mousewheel(event):
    files_canvas.yview_scroll(int(-1*(event.delta/120)), "units")

def _on_shift_mousewheel(event):
    files_canvas.xview_scroll(int(-1*(event.delta/120)), "units")

files_canvas.bind_all("<MouseWheel>", _on_mousewheel)
files_canvas.bind_all("<Shift-MouseWheel>", _on_shift_mousewheel)

# Keyboard bindings
root.bind('<Up>', lambda event: files_canvas.yview_scroll(-1, "units"))
root.bind('<Down>', lambda event: files_canvas.yview_scroll(1, "units"))
root.bind('<Left>', lambda event: files_canvas.xview_scroll(-1, "units"))
root.bind('<Right>', lambda event: files_canvas.xview_scroll(1, "units"))
root.bind('<Prior>', lambda event: files_canvas.yview_scroll(-1, "pages"))
root.bind('<Next>', lambda event: files_canvas.yview_scroll(1, "pages"))

# File 1 treeview
frame1 = tk.LabelFrame(files_scroll_frame, text="First Excel File", background="#7b6cd9", 
                      foreground="white", font=("Poppins", 8, "bold"), width=700)
frame1.pack(padx=5, pady=5, fill="both", expand=True)

tree1_frame = ttk.Frame(frame1, width=1000)
tree1_frame.pack(fill="both", expand=True)

tree1_scroll_y = ttk.Scrollbar(tree1_frame)
tree1_scroll_y.pack(side="right", fill="y")

tree1_scroll_x = ttk.Scrollbar(tree1_frame, orient="horizontal")
tree1_scroll_x.pack(side="bottom", fill="x")

tree1 = ttk.Treeview(tree1_frame, yscrollcommand=tree1_scroll_y.set, xscrollcommand=tree1_scroll_x.set)
tree1.bind("<Button-1>", on_column_click)
tree1.pack(fill="both", expand=True)
tree1["show"] = "headings"

tree1_scroll_y.config(command=tree1.yview)
tree1_scroll_x.config(command=tree1.xview)

# File 2 treeview
frame2 = tk.LabelFrame(files_scroll_frame, text="Second Excel File", background="#7b6cd9", 
                      foreground="white", font=("Poppins", 8, "bold"))
frame2.pack(padx=5, pady=5, fill="both", expand=True)

tree2_frame = ttk.Frame(frame2)
tree2_frame.pack(fill="both", expand=True)

tree2_scroll_y = ttk.Scrollbar(tree2_frame)
tree2_scroll_y.pack(side="right", fill="y")

tree2_scroll_x = ttk.Scrollbar(tree2_frame, orient="horizontal")
tree2_scroll_x.pack(side="bottom", fill="x")

tree2 = ttk.Treeview(tree2_frame, yscrollcommand=tree2_scroll_y.set, xscrollcommand=tree2_scroll_x.set)
tree2.bind("<Button-1>", on_column_click_2)
tree2.pack(fill="both", expand=True)
tree2["show"] = "headings"

tree2_scroll_y.config(command=tree2.yview)
tree2_scroll_x.config(command=tree2.xview)

# ---------------------------- Right Panel (Controls) ---------------------------- #

control_frame = ttk.LabelFrame(main_horizontal_frame, text="Hierarchical Selection", width=300)
control_frame.pack(side="right", fill="y", padx=5, pady=5)
control_frame.pack_propagate(False)

# File 1 hierarchy
file1_hierarchy_frame = ttk.LabelFrame(control_frame, text="File 1 Hierarchy")
file1_hierarchy_frame.grid(row=0, column=0, columnspan=2, pady=5, padx=5, sticky="ew")

# Sheet selection
sheet1_label = ttk.Label(file1_hierarchy_frame, text="-- Sheet:")
sheet1_label.grid(row=0, column=0, sticky="w", pady=5, padx=5)
selected_sheet1 = StringVar(root)
selected_sheet1.set("Sheet1")
sheet1_dropdown = ttk.OptionMenu(file1_hierarchy_frame, selected_sheet1, "Sheet1")
sheet1_dropdown.grid(row=0, column=1, pady=5, padx=5, sticky="ew")
selected_sheet1.trace('w', on_sheet1_selected)

# Table selection
table1_label = ttk.Label(file1_hierarchy_frame, text="    |-- Tables:")
table1_label.grid(row=1, column=0, sticky="w", pady=5, padx=5)
selected_table1 = StringVar(root)
selected_table1.set("Full Sheet")
table1_dropdown = ttk.OptionMenu(file1_hierarchy_frame, selected_table1, "Full Sheet")
table1_dropdown.grid(row=1, column=1, pady=5, padx=5, sticky="ew")
selected_table1.trace('w', on_table1_selected)

# Column selection
column1_label = ttk.Label(file1_hierarchy_frame, text="        |-- Columns:")
column1_label.grid(row=2, column=0, sticky="w", pady=5, padx=5)
selected_col1 = StringVar(root)
selected_col1.set("Select column")
column1_dropdown = ttk.OptionMenu(file1_hierarchy_frame, selected_col1, "Select column")
column1_dropdown.grid(row=2, column=1, pady=5, padx=5, sticky="ew")

# Separator
ttk.Separator(control_frame, orient='horizontal').grid(row=1, column=0, columnspan=2, sticky='ew', pady=10)

# File 2 hierarchy
file2_hierarchy_frame = ttk.LabelFrame(control_frame, text="File 2 Hierarchy")
file2_hierarchy_frame.grid(row=2, column=0, columnspan=2, pady=5, padx=5, sticky="ew")

# Sheet selection
sheet2_label = ttk.Label(file2_hierarchy_frame, text="-- Sheet:")
sheet2_label.grid(row=0, column=0, sticky="w", pady=5, padx=5)
selected_sheet2 = StringVar(root)
selected_sheet2.set("Sheet1")
sheet2_dropdown = ttk.OptionMenu(file2_hierarchy_frame, selected_sheet2, "Sheet1")
sheet2_dropdown.grid(row=0, column=1, pady=5, padx=5, sticky="ew")
selected_sheet2.trace('w', on_sheet2_selected)

# Table selection
table2_label = ttk.Label(file2_hierarchy_frame, text="    |-- Tables:")
table2_label.grid(row=1, column=0, sticky="w", pady=5, padx=5)
selected_table2 = StringVar(root)
selected_table2.set("Full Sheet")
table2_dropdown = ttk.OptionMenu(file2_hierarchy_frame, selected_table2, "Full Sheet")
table2_dropdown.grid(row=1, column=1, pady=5, padx=5, sticky="ew")
selected_table2.trace('w', on_table2_selected)

# Column selection
column2_label = ttk.Label(file2_hierarchy_frame, text="        |-- Columns:")
column2_label.grid(row=2, column=0, sticky="w", pady=5, padx=5)
selected_col2 = StringVar(root)
selected_col2.set("Select column")
column2_dropdown = ttk.OptionMenu(file2_hierarchy_frame, selected_col2, "Select column")
column2_dropdown.grid(row=2, column=1, pady=5, padx=5, sticky="ew")

# Separator
ttk.Separator(control_frame, orient='horizontal').grid(row=3, column=0, columnspan=2, sticky='ew', pady=10)

# Match settings
threshold_frame = ttk.LabelFrame(control_frame, text="Match Settings")
threshold_frame.grid(row=4, column=0, columnspan=2, pady=10, padx=5, sticky="ew")

threshold_label = ttk.Label(threshold_frame, text="Minimum Match %:")
threshold_label.pack(side="left", padx=5)

threshold_var = tk.StringVar(value="80")
threshold_spinbox = ttk.Spinbox(
    threshold_frame,
    from_=0,
    to=100,
    textvariable=threshold_var,
    width=5
)
threshold_spinbox.pack(side="right", padx=5, pady=5)

# Advanced options frame
advanced_options_frame = ttk.LabelFrame(control_frame, text="Advanced Matching Options")
advanced_options_frame.grid(row=5, column=0, columnspan=2, pady=5, padx=5, sticky="ew")

# Algorithm selection
ttk.Label(advanced_options_frame, text="Algorithm:").grid(row=0, column=0, sticky="w", padx=5)
algorithm_var = tk.StringVar(value="Levenshtein")
algorithm_menu = ttk.OptionMenu(
    advanced_options_frame, 
    algorithm_var, 
    "Levenshtein", 
    "Levenshtein", 
   "Jaro-Winkler",  
"Ratio"  
)  
algorithm_menu.grid(row=0, column=1, sticky="ew", padx=5)

ttk.Label(advanced_options_frame, text="Data Preprocessing:").grid(row=1, column=0, sticky="w", padx=5)

trim_whitespace_var = tk.BooleanVar(value=True)
trim_whitespace_cb = ttk.Checkbutton(
advanced_options_frame,
text="Trim whitespace",
variable=trim_whitespace_var
)

trim_whitespace_cb.grid(row=2, column=0, columnspan=2, sticky="w", padx=5)

normalize_case_var = tk.BooleanVar(value=True)
normalize_case_cb = ttk.Checkbutton(
advanced_options_frame,
text="Normalize case",
variable=normalize_case_var
)
normalize_case_cb.grid(row=3, column=0, columnspan=2, sticky="w", padx=5)

ignore_punct_var = tk.BooleanVar(value=True)
ignore_punct_cb = ttk.Checkbutton(
advanced_options_frame,
text="Ignore punctuation",
variable=ignore_punct_var
)
ignore_punct_cb.grid(row=4, column=0, columnspan=2, sticky="w", padx=5)

case_sensitive_var = tk.BooleanVar(value=False)
case_sensitive_cb = ttk.Checkbutton(
advanced_options_frame,
text="Case sensitive",
variable=case_sensitive_var
)
case_sensitive_cb.grid(row=5, column=0, columnspan=2, sticky="w", padx=5)


advanced_toggle_btn = ttk.Button(
control_frame,
text="Show Advanced Options",
command=show_advanced_options
)
advanced_toggle_btn.grid(row=6, column=0, columnspan=2, pady=5)

match_btn = round_button(control_frame, radius=25, fill="#7b6cd9", font=("Poppins", 9, "bold"),
width=170, text="Start matching", command=start_matching)
match_btn.grid(row=7, column=0, columnspan=2, pady=20, padx=10, sticky="nsew")

status_label = tk.Label(control_frame, text="Ready", bd=1, relief="sunken", anchor="w")
status_label.grid(row=8, column=0, columnspan=2, pady=5, padx=5, sticky="ew")

progress_var = tk.DoubleVar()
result_filename = ""
progress_window = None

root.mainloop()