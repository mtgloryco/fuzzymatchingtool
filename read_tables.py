import pandas as pd
from openpyxl import load_workbook

file_path = 'tables.xlsx'
wb = load_workbook(filename=file_path)


table_info = []
for sheet in wb.worksheets:
    for table in sheet.tables.values():
        table_info.append((table.name, sheet.title, table.ref))


print("List of tables:")
for idx, (name, sheet, ref) in enumerate(table_info):
    print(f"{idx + 1}: {name} (Sheet: {sheet}, Range: {ref})")


choice = int(input("Enter the number of the table to print: ")) - 1
if 0 <= choice < len(table_info):
    name, sheet_name, ref = table_info[choice]
    ws = wb[sheet_name]
    data = ws[ref]
    rows = [[cell.value for cell in row] for row in data]
    df = pd.DataFrame(rows[1:], columns=rows[0])
    print(f"\nData from table '{name}':")
    print(df)
else:
    print("Invalid selection.")