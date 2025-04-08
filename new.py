import openpyxl as op
import pandas as pd
from fuzzywuzzy import process
from tkinter import filedialog, messagebox, StringVar
import collections
import tkinter as tk
from tkinter import ttk
import os
from datetime import datetime

class ExcelFuzzyMatcher:
    def __init__(self, master):
        self.master = master
        self.master.title("Excel Fuzzy Matching")

        self.style = ttk.Style()
        self.style.theme_use("default")
        self.style.configure("Treeview.Heading", foreground="red", font=('Helvetica', 10, 'bold'))

        self.tree = ttk.Treeview(self.master)
        self.tree.pack()

        self.file1_path = ""
        self.file2_path = ""
        self.options = []
        self.selected_col = StringVar()

        self.setup_widgets()

    def setup_widgets(self):
        self.file1_label = tk.Label(self.master, text="Select first Excel file", width=50)
        self.file1_label.pack(pady=5)
        tk.Button(self.master, text="Choose File 1", command=self.select_file1).pack(pady=5)

        self.file2_label = tk.Label(self.master, text="Select second Excel file", width=50)
        self.file2_label.pack(pady=5)
        tk.Button(self.master, text="Choose File 2", command=self.select_file2).pack(pady=5)

        self.match_btn = tk.Button(self.master, text="Start Matching", command=self.start_matching, bg="green", fg="white")
        self.match_btn.pack(pady=10)

    def drop_down_menu(self, columns, file_label):
        self.options = columns
        self.selected_col.set(self.options[0])

        drop_down_label = tk.Label(self.master, text=f"Columns for {file_label} ")
        drop_down_label.pack()
        drop_down = tk.OptionMenu(self.master, self.selected_col, *self.options)
        drop_down.pack()

    def read_columns(self, file, file_label):
        try:
            sheet = pd.read_excel(file)
            cols = sheet.columns.tolist()
            self.tree.config(columns=cols)
            self.drop_down_menu(cols, file_label)
            for col in cols:
                self.tree.heading(col, text=col)
                self.tree.column(col, anchor='center')
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read {file_label}: {e}")

    def select_file1(self):
        self.file1_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if self.file1_path:
            self.file1_label.config(text=f"Selected: {os.path.basename(self.file1_path)}")
            self.read_columns(self.file1_path, "File 1")

    def select_file2(self):
        self.file2_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if self.file2_path:
            self.file2_label.config(text=f"Selected: {os.path.basename(self.file2_path)}")
            self.read_columns(self.file2_path, "File 2")

    def start_matching(self):
        try:
            if not self.file1_path or not self.file2_path:
                raise ValueError("Both Excel files must be selected.")

            col_name = self.selected_col.get()
            if col_name not in self.options:
                raise ValueError("Selected column is not valid.")

            file1 = op.load_workbook(self.file1_path)
            file2 = op.load_workbook(self.file2_path)
            sheet1 = file1.active
            sheet2 = file2.active

            index_of_item = self.options.index(col_name)
            email2 = [r[index_of_item] for r in sheet2.iter_rows(values_only=True) if r[index_of_item]]

            result = op.Workbook()
            result_sheet = result.active
            result_sheet.title = "Result"
            result_sheet.append(["Email1", "Matched Email", "Match Score"])

            res = collections.defaultdict(list)
            row_tracker = 1

            for t in sheet1.iter_rows(values_only=True):
                email = t[index_of_item]
                if email:
                    match = process.extract(email, email2, limit=1)[0]
                    res[email] = match
                    result_sheet.append([email, match[0], match[1]])
                    row_tracker += 1

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            result_filename = f"Result_{timestamp}.xlsx"
            result.save(result_filename)
            messagebox.showinfo("Success", f"Matching complete! Results saved as '{result_filename}'.")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")


if __name__ == '__main__':
    root = tk.Tk()
    app = ExcelFuzzyMatcher(root)
    root.mainloop()
